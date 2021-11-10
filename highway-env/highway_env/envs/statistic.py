import numpy as np
import openpyxl
from openpyxl import Workbook
import copy
import os
import json
from typing import List, Tuple, Union, Optional, Callable
import gym
import random
from gym import Wrapper
from gym.utils import seeding
from gym.envs.registration import register
from highway_env import utils

from highway_env.road.road import Road, RoadNetwork, LaneIndex
from highway_env.vehicle.controller import ControlledVehicle
from highway_env.envs.common.action import action_factory, Action, DiscreteMetaAction, ActionType
from highway_env.envs.common.observation import observation_factory, ObservationType
from highway_env.envs.common.finite_mdp import finite_mdp
from highway_env.envs.common.graphics import EnvViewer
from highway_env.vehicle.behavior import IDMVehicle, LinearVehicle
from highway_env.vehicle.controller import MDPVehicle
from highway_env.vehicle.kinematics import Vehicle
from highway_env.envs.json import Json

LaneIndex = Tuple[str, str, int]
Action = Union[int, np.ndarray]

class Statistics(gym.Env):
    
    def liste(self, ego_vehicle : Vehicle, action = Action, distance_algo = int):
        """
        overtakeCount: calls overtaking function in Behaviuor.py
        unsafe: calls safedistance function in Behaviuor.py
        right: variable set to 1 only if the vehicle is in the rightmost lane, variable that will be multiplied with RIGHT_LANE_REWARD
        start: varibale for save the position of the vehicle at the beginning of the simulation
        """
        
        overtakeCount, right_id = IDMVehicle.rightovertaking(self, ego_vehicle = self.vehicle, lane_index = self.vehicle.lane_index) 

        idrig = 0
        if overtakeCount != 0 and len(self.list_rightID) != 0:
            for rightID in self.list_rightID:
                if rightID == right_id:
                    idrig += 1
            if idrig == 0:
                self.list_rightID.append(right_id)
        if overtakeCount != 0 and len(self.list_rightID) == 0:
            self.list_rightID.append(right_id)  
        

        leftOvertake, vehicle_id = IDMVehicle.leftovertaking(self, ego_vehicle = self.vehicle, lane_index = self.vehicle.lane_index) #check if the vehicle passed another one on the left

        lane = 0.0
        if self.vehicle.lane_index[2] == 2:
            lane = 0.0
        if self.vehicle.lane_index[2] == 1:
            lane = 0.5
        if self.vehicle.lane_index[2] == 0:
            lane = 1.0

        if leftOvertake != 0 and len(self.list_leftID) != 0:
            for leftID in self.list_leftID:
                if leftID == vehicle_id:
                    leftOvertake = 0
            if leftOvertake != 0:
                self.list_leftID.append(vehicle_id)
        elif leftOvertake != 0 and len(self.list_leftID) == 0:
            self.list_leftID.append(vehicle_id)
        
        if distance_algo == 0:
            ttc = IDMVehicle.safedistance(self, ego_vehicle = self.vehicle, lane_index = self.vehicle.lane_index)             
            tth, id_vehicle = IDMVehicle.hazardous(self, ego_vehicle = self.vehicle, lane_index = self.vehicle.lane_index) 
            idhaz = 0
            l = len(self.list_hlc)

            if ttc < 3:
                unsafe = (1- (ttc / 3)) 
            else:
                ttc = 3
                unsafe = 0
            if tth < 15:
                hazardousLC = (1- (tth / 15)) 
                if l > 0:
                    for hazID in self.list_hlc:
                        if hazID == id_vehicle:
                            idhaz += 1
                    if idhaz == 0:
                        self.list_hlc.append(id_vehicle)
                else:
                    self.list_hlc.append(id_vehicle)  
            else:
                tth = 15
                hazardousLC = 0  
        else:
            ttc = IDMVehicle.safedistance_TTC(self, ego_vehicle = self.vehicle, lane_index = self.vehicle.lane_index)            
            tth, id_vehicle = IDMVehicle.hazardous_lane_change_TTC(self, ego_vehicle = self.vehicle, lane_index = self.vehicle.lane_index) 

            idhaz = 0
            l = len(self.list_hlc)

            if ttc < 3:
                unsafe = (1- (ttc / 3)) 
            else:
                ttc = 3
                unsafe = 0
            if tth < 15:
                hazardousLC = (1- (tth / 15)) 
                if l > 0:
                    for hazID in self.list_hlc:
                        if hazID == id_vehicle:
                            idhaz += 1
                    if idhaz == 0:
                        self.list_hlc.append(id_vehicle)
                else:
                    self.list_hlc.append(id_vehicle)  
            else:
                tth = 15
                hazardousLC = 0                          
        
        if self.vehicle.speed < self.config["limit_speed"]-5: #vel<25
            vel = 1.0
        elif self.vehicle.speed >= self.config["limit_speed"]-5 and self.vehicle.speed <= self.config["limit_speed"] -1: #25<vel<29
            vel = (-0.2*self.vehicle.speed) + self.config["limit_speed"]/5 
        elif self.vehicle.speed > self.config["limit_speed"] and self.vehicle.speed <= self.config["limit_speed"] + 2: #30<vel<32
            vel = (0.50*self.vehicle.speed) - self.config["limit_speed"]/2 
        elif self.vehicle.speed > self.config["limit_speed"] + 2: #vel>32
            vel = (0.50*self.config["limit_speed"] + 5) - self.config["limit_speed"]/2 
        else: #29<vel<30
            vel = 0.0

        if self.vehicle.speed >= self.config["limit_speed"] - 1 and self.vehicle.speed <= self.config["limit_speed"]: #vel > 29 and vel < 30
            vel_pos = 1.0
        else:
            vel_pos = 0.0

        acc = 0
        dec = 0
        if self.vehicle.action['acceleration'] >= 0.001:
            acc = round(self.vehicle.action['acceleration'],3)
        elif self.vehicle.action['acceleration'] <= -0.001:
            dec = round(abs(self.vehicle.action['acceleration']),3)
        else:
            acc = 0
            dec = 0 

        steering = abs(10*self.vehicle.action['steering'])
          
    
        self.list_vel.append(self.vehicle.speed)
        if len(self.list_vel) == 1:
            start = self.vehicle.position[0]
                
        #self.list_over.append(overtakeCount)
        self.list_unsafe.append(ttc)
        self.list_right.append(self.vehicle.lane_index[2])
        self.list_acc.append(acc)
        self.list_dec.append(dec)
        self.list_nChangeLane.append(action)
        self.list_str.append(steering)

        return overtakeCount, unsafe, vel, leftOvertake, acc, dec, hazardousLC, steering, vel_pos, lane
    
    def statistics(self, distance_algo = int):
        """Return the statistics of my environment"""
        
        finish = 0
        space_trav = 0
        s_t = 0.0
        r_o = 0
        l_o = 0
        avg_c = 0
        std_c = 0
        avg_dsl = 0
        std_dsl = 0
        avg_loaf = 0
        n_lc = 0
        hlc = 0
        crush = False

        n_step = len(self.list_vel)
        som = 0
        i = 0

        """
        Crush
        """
        if self.vehicle.crashed:
            crush = True

        """
        Km driven
        """
        finish = self.vehicle.position[0]
        space_trav = finish - self.start
        s_t = space_trav/2000

        """
        loop to see how many times the vehicle overtakes another vehicle on the right
        """
        r_o = len(self.list_rightID)
        del self.list_rightID[0:n_step]
        #del self.list_over[0:n_step] 
        i = 0         

        """
        number of left-overtake
        """
        l_o = len(self.list_leftID)
        del self.list_leftID[0:n_step] 

        """
        AVG TTC
        """
        while i < n_step:
            som += self.list_unsafe[i]            
            self.list_stdttc.append(self.list_unsafe[i])
            i += 1
        avg_c = som / n_step
        som = 0
        i = 0

        """
        STD TTC
        """
        while i < n_step:
            som += pow((self.list_unsafe[i] - avg_c), 2)
            i += 1
        std_c = pow((som/n_step), 1/2)
        del self.list_unsafe[0:n_step]
        som = 0
        i = 0

        """
        AVG Delta speed Limit
        """
        threshold = self.config["limit_speed"]        
        while i < n_step:
            som += threshold - self.list_vel[i]
            self.list_stddsl.append(threshold - self.list_vel[i])
            i += 1
        avg_dsl = som / n_step
        som = 0
        i = 0

        """
        STD Delta Speed Limit
        """
        while i < n_step:
            som += pow(((threshold - self.list_vel[i]) - avg_dsl), 2)
            i += 1
        std_dsl = pow((som/n_step), 1/2)
        del self.list_vel[0:n_step]
        som = 0
        i = 0

        """
        AVG Lane occupacy function
        """       
        while i < n_step:
            som += self.list_right[i]
            i += 1
        avg_loaf = som / n_step
        del self.list_right[0:n_step]
        som = 0
        i = 0

        """
        loop see how many times the vehicle changes lanes
        """
        while i < n_step - 1:
            if self.list_nChangeLane[i] == 0 or self.list_nChangeLane[i] == 2: #"0": left lane, "2": right lane
                    n_lc += 1  
            i += 1
        del self.list_nChangeLane[0:n_step]
        i = 0

        """
        numbers of hazardous lane change
        """
        hlc = len(self.list_hlc) 
        del self.list_hlc[0:len(self.list_hlc)]    

        """
        AVG acceleration
        """       
        while i < n_step:
            som += self.list_acc[i]            
            self.list_stdacc.append(self.list_acc[i])
            i += 1
        avg_acc = som / n_step
        som = 0
        i = 0

        """
        STD acceleration
        """
        while i < n_step:
            som += pow((self.list_acc[i] - avg_acc), 2)
            i += 1
        std_acc = pow((som/n_step), 1/2)
        del self.list_acc[0:n_step]
        som = 0
        i = 0

        """
        AVG deceleration
        """      
        while i < n_step:
            som += self.list_dec[i]
            self.list_stddec.append(self.list_dec[i])
            i += 1
        avg_dec = som / n_step     
        som = 0
        i = 0

        """
        STD deceleration
        """
        while i < n_step:
            som += pow((self.list_dec[i] - avg_dec), 2)
            i += 1
        std_dec = pow((som/n_step), 1/2) 
        del self.list_dec[0:n_step]  
        som = 0
        i = 0    

        """
        AVG steering   
        """ 
  
        while i < n_step:
            som += self.list_str[i]
            self.list_stdstr.append(self.list_str[i])
            i += 1
        avg_str = som / n_step    
        som = 0
        i = 0
        
        """
        STD steering
        """
        while i < n_step:
            som += pow((self.list_str[i] - avg_str), 2)
            i += 1
        std_str = pow((som/n_step), 1/2)
        del self.list_str[0:n_step]    
        som = 0
        i = 0          
        
        
        """
        save the parameters in list_evaluate
        """
        self.list_evaluate.append(crush)
        self.list_evaluate.append(s_t)
        self.list_evaluate.append(r_o)
        self.list_evaluate.append(l_o)
        self.list_evaluate.append(avg_c)
        self.list_evaluate.append(std_c)
        self.list_evaluate.append(avg_dsl)
        self.list_evaluate.append(std_dsl)
        self.list_evaluate.append(avg_loaf)
        self.list_evaluate.append(n_lc)
        self.list_evaluate.append(hlc)
        self.list_evaluate.append(avg_acc)
        self.list_evaluate.append(std_acc)
        self.list_evaluate.append(avg_dec)
        self.list_evaluate.append(std_dec)
        self.list_evaluate.append(avg_str)
        self.list_evaluate.append(std_str)
        ev_ep = Statistics.evaluateEP(self, self.list_evaluate)
        self.list_evaluate.append(ev_ep)
        self.lista.append(1) #episode counter

        """
        the length of "list" is compared with the number of episodes to call "evaluate" only when all the episodes have been simulated
        """
        if len(self.lista) == self.config["number_episodes"]: 
            Statistics.printEpisodes(self, self.list_evaluate, self.config["SafeDistance"])
            del self.list_evaluate[0:self.config["number_episodes"]*18] #18: number of metrics
            del self.lista[0:self.config["number_episodes"]]
    

    def printEpisodes(self, list_evaluate, distance_algo = int):
        """Method that evaluate the episodes"""
        
        excel_document = openpyxl.load_workbook(os.environ.get('HIGHWAY_ENV_PATH') + 'EvaluationPolicies.xlsx') #load file excel esistente
        #according to the value of "env" I overwrite the corresponding sheet
        if self.config["env"] == 0:
            foglio = excel_document.get_sheet_by_name('RL Policy') 
        elif self.config["env"] == 1:
            foglio = excel_document.get_sheet_by_name('Random Policy')
        elif self.config["env"] == 2:
            foglio = excel_document.get_sheet_by_name('Heuristic Policy')

        foglio["A1"] = "Ep"
        foglio["B64"] = foglio["B1"] = "CR"
        foglio["C64"] = foglio["C1"] = "Km"
        foglio["D64"] = foglio["D1"] = "RO"
        foglio["E64"] = foglio["E1"] = "LO"
        if distance_algo == 0:
            foglio["F64"] = foglio["F1"] = "AVG DTC"
            foglio["G64"] = foglio["G1"] = "STD DTC"
        else:
            foglio["F64"] = foglio["F1"] = "AVG TTC"
            foglio["G64"] = foglio["G1"] = "STD TTC"
        foglio["H64"] = foglio["H1"] = "AVG DSL"
        foglio["I64"] = foglio["I1"] = "STD DSL"
        foglio["J64"] = foglio["J1"] = "AVG LOAF"
        foglio["K64"] = foglio["K1"] = "LC"
        foglio["L64"] = foglio["L1"] = "HAZ"
        foglio["M64"] = foglio["M1"] = "AVG ACC"
        foglio["N64"] = foglio["N1"] = "STD ACC"
        foglio["O64"] = foglio["O1"] = "AVG DEC"
        foglio["P64"] = foglio["P1"] = "STD DEC"
        foglio["Q64"] = foglio["Q1"] = "AVG STE"
        foglio["R64"] = foglio["R1"] = "STD STE"
        foglio["S1"] = "EV EP"

        foglio["Z1"] = "EP: episode #"
        foglio["Z2"] = "CR: crush"
        foglio["Z3"] = "Km: # Km"
        foglio["Z4"] = "RO: #right overtake"
        foglio["Z5"] = "LO: # left overtake"
        if distance_algo == 0:
            foglio["Z6"] = "AVG DTC: average distance to collision"
            foglio["Z7"] = "STD DTC: standard deviation distance to collision"
        else:
            foglio["Z6"] = "AVG TTC: average TTC"
            foglio["Z7"] = "STD TTC: standard deviation TTC"
        foglio["Z8"] = "AVG DSL: average module delta speed limit"
        foglio["Z9"] = "STD DSL: standard deviatio delta speed limit"
        foglio["Z10"] = "AVG LOAF: average lane occupacy assestment function"
        foglio["Z11"] = "LC: # lane changes"
        foglio["Z12"] = "HAZ: # hazardous lane change"
        foglio["Z13"] = "AVG ACC: average acceleration"
        foglio["Z14"] = "STD ACC: standard deviation acceleration"
        foglio["Z15"] = "AVG DEC: average deceleration"
        foglio["Z16"] = "STD DEC: standard deviation deceleration"
        foglio["Z17"] = "AVG STE: average stering"
        foglio["Z18"] = "STD STE: standard deviation stering"
        foglio["Z19"] = "EV EP: evaluation episode"

        foglio["A65"] = "1h Ep"
        foglio["A67"] = "Quality"
        foglio["A68"] = "Quality %"

        if self.config["env"] == 0:
            foglio["Z28"] = "Right_Lane_Reward"
            foglio["Z29"] = "High_Speed_Reward"
            foglio["Z30"] = "Speed_Reward"
            foglio["Z31"] = "Right_Overtaking_Reward"
            foglio["Z32"] = "Unsafe_Distance_Reward"
            foglio["Z33"] = "Left_Overtaking_Reward"
            foglio["Z34"] = "Long_Accel_Reward"
            foglio["Z35"] = "Long_Decel_Reward"
            foglio["Z36"] = "Hazardous_Lane_Change_Reward"
            foglio["Z37"] = "Steering_Reward"
            foglio["Z38"] = "Collision_reward"

            foglio["AD28"] = self.config["Right_Lane_Reward"]
            foglio["AD29"] = self.config["High_Speed_Reward"]
            foglio["AD30"] = self.config["Speed_Reward"]
            foglio["AD31"] = self.config["Right_Overtaking_Reward"]
            foglio["AD32"] = self.config["Unsafe_Distance_Reward"]
            foglio["AD33"] = self.config["Left_Overtaking_Reward"]
            foglio["AD34"] = self.config["Long_Accel_Reward"]
            foglio["AD35"] = self.config["Long_Decel_Reward"]
            foglio["AD36"] = self.config["Hazardous_Lane_Change_Reward"]
            foglio["AD37"] = self.config["Steering_Reward"]
            foglio["AD38"] = self.config["collision_reward"]


        i = 0

        """
        loop to enter episode numbers in the first column
        """
        for r in range (2, self.config["number_episodes"]+2):
            c = 1
            i += 1
            foglio.cell(row = r, column = c, value = i)

        i = 0
        
        """
        loop to fill the cells of the excel sheet
        """
        for r in range (2, self.config["number_episodes"]+2):
            for c in range(2 , 20): 
                foglio.cell(row = r, column = c, value = round(self.list_evaluate[i],1))
                if i < self.config["number_episodes"]*18: #17: number of metrics
                    i += 1

        i = 0
        score = 0
        perc = 0
        score, perc = Statistics.evaluatePolicy(self)
        r = 65
        for c in range(2 , 19): #16: number of the last row
            foglio.cell(row = r, column = c, value = round(self.list_hour[i],1))
            if i < 19: #17: number of metrics
                    i += 1
        i = 0
        r = 67
        for c in range(2 , 19): #16: number of the last row
            foglio.cell(row = r, column = c, value = self.list_perc[i])
            if i < 19: #17: number of metrics
                    i += 1

        i = 0
        r = 68
        for c in range(2 , 19): #16: number of the last row
            foglio.cell(row = r, column = c, value = self.list_p[i])
            if i < 19: #17: number of metrics
                    i += 1            
            
        foglio["A70"] = "The policy evaluation is:" 
        foglio["D70"] = score
        foglio["E70"] = "out of"
        foglio["F70"] = perc

        del self.list_hour[0:len(self.list_hour)]
        del self.list_perc[0:len(self.list_perc)]
        del self.list_p[0:len(self.list_p)]

        """
        save the excel sheet
        """
        excel_document.save(filename = os.environ.get('HIGHWAY_ENV_PATH') + "EvaluationPolicies.xlsx")

    
    def evaluatePolicy(self):
        """Method that evaluate the Policy"""
        data = Json.config_json_weight(self)        
        data1 = Json.config_json_max(self)
        perc_crush = 0.0  
        perc_space = 0.0 
        perc_r_over = 0.0   
        perc_l_over = 0.0   
        perc_avg_t = 0.0   
        perc_std_t = 0.0   
        perc_avg_d = 0.0  
        perc_std_d = 0.0  
        perc_avg_l = 0.0  
        perc_lane = 0.0  
        perc_hazardous = 0.0 
        perc_avg_acc = 0.0 
        perc_std_acc = 0.0   
        perc_avg_dec = 0.0   
        perc_std_dec = 0.0   
        perc_avg_str = 0.0  
        perc_std_str = 0.0 

        totWeight = data['weight_crush'] + data['weight_km'] + data['weight_right_overtake'] + data['weight_left_overtake'] + data['weight_AVG_TTC'] + data['weight_STD_TTC'] + data['weight_AVG_DSL'] + data['weight_STD_DSL'] + data['weight_AVG_LOAF'] + data['weight_number_lane_change'] + data['weight_hazardous'] + data['weight_AVG_ACC'] + data['weight_STD_ACC'] + data['weight_AVG_DEC'] + data['weight_STD_DEC'] + data['weight_AVG_STER'] + data['weight_STD_STER'] 
        
        """
        Crush in 1 hour
        """
        i=0
        crushed = 0
        while (i<self.config["number_episodes"]*18):
            if (self.list_evaluate[i]==True):
                crushed += 1
            i += 18
        if (crushed > data1['max_crush_per_h']):
            perc_crush = 0
        else:
            perc_crush = (1 - (crushed/data1['max_crush_per_h']))*data['weight_crush']
        self.list_hour.append(crushed)
        self.list_perc.append(str(int(perc_crush)) +"/" +str(data['weight_crush']))   
        if  perc_crush < data['weight_crush']:
            p = (perc_crush/data['weight_crush'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")    
        """
        Km in 1 hour
        """
        i=1
        space = 0
        while (i<self.config["number_episodes"]*18):
            space += self.list_evaluate[i]
            i += 18
        if (space > data1['max_km_per_h']):
            perc_space = data['weight_km']
        else:
            perc_space = (space/data1['max_km_per_h'])*data['weight_km']
        self.list_hour.append(space)
        self.list_perc.append(str(int(perc_space)) +"/" +str(data['weight_km']))
        if  perc_space < data['weight_km']:
            p = (perc_space/data['weight_km'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")    
        """
        Right overtake in 1 hour
        """
        i=2
        r_over = 0
        while (i<self.config["number_episodes"]*18):
            r_over += self.list_evaluate[i]
            i +=18
        if (r_over > data1['max_right_overtake_per_h']):
            perc_r_over = 0 
        else:
            perc_r_over = (1 - (r_over/data1['max_right_overtake_per_h']))*data['weight_right_overtake']      
        self.list_hour.append(r_over)
        self.list_perc.append(str(int(perc_r_over)) +"/" +str(data['weight_right_overtake'])) 
        if  perc_r_over < data['weight_right_overtake']:
            p = (perc_r_over/data['weight_right_overtake'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")   
        """
        Left overtake in 1 hour
        """
        i=3
        l_over = 0
        while (i<self.config["number_episodes"]*18):
            l_over += self.list_evaluate[i]
            i +=18
        if (l_over > data1['max_left_overtake_per_h']):
            perc_l_over = data['weight_left_overtake'] 
        else:
            perc_l_over = (l_over/data1['max_left_overtake_per_h'])*data['weight_left_overtake']  
        self.list_hour.append(l_over)
        self.list_perc.append(str(int(perc_l_over)) +"/" +str(data['weight_left_overtake'])) 
        if  perc_l_over < data['weight_left_overtake']:
            p = (perc_l_over/data['weight_left_overtake'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%") 
        """
        AVG TTC in 1 hour
        """
        i=4
        avg_t = 0
        while (i<self.config["number_episodes"]*18):
            avg_t += self.list_evaluate[i]
            i +=18
        avg_t = avg_t/self.config["number_episodes"]
        if (avg_t > data1['max_AVG_TTC_per_h']):
            perc_avg_t = data['weight_AVG_TTC']
        else:
            perc_avg_t = (avg_t/data1['max_AVG_TTC_per_h'])*data['weight_AVG_TTC']      
        self.list_hour.append(avg_t)
        self.list_perc.append(str(int(perc_avg_t)) +"/" +str(data['weight_AVG_TTC'])) 
        if  perc_avg_t < data['weight_AVG_TTC']:
            p = (perc_avg_t/data['weight_AVG_TTC'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%") 
        """
        STD TTC in 1 hour
        """
        i=0
        som_ttc = 0
        while (i < len(self.list_stdttc)):
            som_ttc += pow((self.list_stdttc[i] - avg_t), 2)
            i += 1
        std_t = pow((som_ttc/len(self.list_stdttc)), 1/2)
        if (std_t > data1['max_STD_TTC_per_h']):
            perc_std_t = 0 
        else: 
            perc_std_t = (1 - (std_t/data1['max_STD_TTC_per_h']))*data['weight_STD_TTC'] 
        del self.list_stdttc[0:len(self.list_stdttc)]   
        self.list_hour.append(std_t)
        self.list_perc.append(str(int(perc_std_t)) +"/" +str(data['weight_STD_TTC']))    
        if  perc_std_t < data['weight_STD_TTC']:
            p = (perc_std_t/data['weight_STD_TTC'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%") 
        """
        AVG delta speed limit in 1 hour
        """
        i=6
        avg_d = 0
        while (i<self.config["number_episodes"]*18):
            avg_d += self.list_evaluate[i]
            i +=18
        avg_d = avg_d/self.config["number_episodes"]
        if (avg_d > data1['max_AVG_DSL_per_h']):
            perc_avg_d = 0
        else: 
            perc_avg_d = (1 - (avg_d/data1['max_AVG_DSL_per_h']))*data['weight_AVG_DSL'] 
        self.list_hour.append(avg_d)
        self.list_perc.append(str(int(perc_avg_d)) +"/" +str(data['weight_AVG_DSL'])) 
        if  perc_avg_d < data['weight_AVG_DSL']:
            p = (perc_avg_d/data['weight_AVG_DSL'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")     
        """
        STD delta speed limit in 1 hour
        """
        i=0
        som_dsl = 0
        while (i < len(self.list_stddsl)):
            som_dsl += pow((self.list_stddsl[i] - avg_d), 2)
            i += 1
        std_d = pow((som_dsl/len(self.list_stddsl)), 1/2)
        if (std_d > data1['max_STD_DSL_per_h']):
            perc_std_d = 0 
        else:
            perc_std_d = (1 - (std_d/data1['max_STD_DSL_per_h']))*data['weight_STD_DSL'] 
        del self.list_stddsl[0:len(self.list_stddsl)]
        self.list_hour.append(std_d)
        self.list_perc.append(str(int(perc_std_d)) +"/" +str(data['weight_STD_DSL']))  
        if  perc_std_d < data['weight_STD_DSL']:
            p = (perc_std_d/data['weight_STD_DSL'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")     
        """
        AVG lane occupacy function
        """
        i=8
        avg_l = 0
        while (i<self.config["number_episodes"]*18):
            avg_l += self.list_evaluate[i]
            i +=18
        avg_l = avg_l/self.config["number_episodes"]
        perc_avg_l = (avg_l/data1['max_AVG_LOAF_per_h'])*data['weight_AVG_LOAF']   
        self.list_hour.append(avg_l)
        self.list_perc.append(str(int(perc_avg_l)) +"/" +str(data['weight_AVG_LOAF']))  
        if  perc_avg_l < data['weight_AVG_LOAF']:
            p = (perc_avg_l/data['weight_AVG_LOAF'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")   
        """
        Lanes change in 1 hour
        """
        i=9
        lane = 0
        while (i<self.config["number_episodes"]*18):
            lane += self.list_evaluate[i]
            i +=18  
        if (lane > data1['max_number_lane_change_per_h']):
            perc_lane = 0
        else:
            perc_lane = data['weight_number_lane_change'] 
        self.list_hour.append(round(lane,0))
        self.list_perc.append(str(int(perc_lane)) +"/" +str(data['weight_number_lane_change'])) 
        if  perc_lane < data['weight_number_lane_change']:
            p = (perc_lane/data['weight_number_lane_change'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")   
        """
        Hazardous lane change in 1 h
        """
        i=10
        hazardous = 0
        while (i<self.config["number_episodes"]*18):
            hazardous += self.list_evaluate[i]
            i +=18
        perc_hazardous = (1- (hazardous/data1['max_hazardous_per_h']))*data['weight_hazardous']  
        if (hazardous > data1['max_hazardous_per_h']):
            perc_hazardous = 0
        self.list_hour.append(hazardous)
        self.list_perc.append(str(int(perc_hazardous)) +"/" +str(data['weight_hazardous']))   
        if  perc_hazardous < data['weight_hazardous']:
            p = (perc_hazardous/data['weight_hazardous'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%") 
        """
        AVG acceleration in 1 h
        """
        i=11
        avg_acc = 0
        while (i<self.config["number_episodes"]*18):
            avg_acc += self.list_evaluate[i]
            i +=18
        avg_acc = avg_acc/self.config["number_episodes"] 
        if (avg_acc > data1['max_AVG_ACC_per_h']):
            perc_avg_acc = 0 
        else:
            perc_avg_acc = (1 - (avg_acc/data1['max_AVG_ACC_per_h']))*data['weight_AVG_ACC'] 
        self.list_hour.append(avg_acc)
        self.list_perc.append(str(int(perc_avg_acc)) +"/" +str(data['weight_AVG_ACC']))  
        if  perc_avg_acc < data['weight_AVG_ACC']:
            p = (perc_avg_acc/data['weight_AVG_ACC'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")   
        """
        STD acceleration in 1 h
        """
        i=0
        som_acc = 0
        while (i < len(self.list_stdacc)):
            som_acc += pow((self.list_stdacc[i] - avg_acc), 2)
            i += 1
        std_acc = pow((som_acc/len(self.list_stdacc)), 1/2)
        if (std_acc > data1['max_STD_ACC_per_h']):
            perc_std_acc = 0 
        else: 
            perc_std_acc = (1 - (std_acc/data1['max_STD_ACC_per_h']))*data['weight_STD_ACC']         
        del self.list_stdacc[0:len(self.list_stdacc)] 
        self.list_hour.append(std_acc)
        self.list_perc.append(str(int(perc_std_acc)) +"/" +str(data['weight_STD_ACC']))  
        if  perc_std_acc < data['weight_STD_ACC']:
            p = (perc_std_acc/data['weight_STD_ACC'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%") 
        """
        AVG deceleration in 1 h
        """
        i=13
        avg_dec = 0
        while (i<self.config["number_episodes"]*18):
            avg_dec += self.list_evaluate[i]
            i +=18
        avg_dec = avg_dec/self.config["number_episodes"] 
        if (avg_dec > data1['max_AVG_DEC_per_h']):
            perc_avg_acc = 0  
        else:
            perc_avg_dec = (1- (avg_dec/data1['max_AVG_DEC_per_h']))*data['weight_AVG_DEC'] 
        self.list_hour.append(avg_dec)
        self.list_perc.append(str(int(perc_avg_dec)) +"/" +str(data['weight_AVG_DEC']))   
        if  perc_avg_dec < data['weight_AVG_DEC']:
            p = (perc_avg_dec/data['weight_AVG_DEC'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")   
        """
        STD deceleration 1 hour
        """
        i=0
        som_dec = 0
        while (i < len(self.list_stddec)):
            som_dec += pow((self.list_stddec[i] - avg_dec), 2)
            i += 1
        std_dec = pow((som_dec/len(self.list_stddec)), 1/2)
        if (std_dec > data1['max_STD_DEC_per_h']):
            perc_std_dec = 0 
        else:
            perc_std_dec = (1 - (std_dec/data1['max_STD_DEC_per_h']))*data['weight_STD_DEC']  
        del self.list_stddec[0:len(self.list_stddec)]
        self.list_hour.append(std_dec)
        self.list_perc.append(str(int(perc_std_dec)) +"/" +str(data['weight_STD_DEC']))   
        if  perc_std_dec < data['weight_STD_DEC']:
            p = (perc_std_dec/data['weight_STD_DEC'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%") 
        """
        AVG steering in 1 hour
        """       
        i=15
        avg_str = 0
        while (i<self.config["number_episodes"]*18):
            avg_str += self.list_evaluate[i]
            i +=18
        avg_str = avg_str/self.config["number_episodes"]
        if (avg_str > data1['max_AVG_STER_per_h']):
            perc_avg_str = 0
        else: 
            perc_avg_str = (1 - (avg_str/data1['max_AVG_STER_per_h']))*data['weight_AVG_STER']   
        self.list_hour.append(avg_str)
        self.list_perc.append(str(int(perc_avg_str)) +"/" +str(data['weight_AVG_STER']))   
        if  perc_avg_str < data['weight_AVG_STER']:
            p = (perc_avg_str/data['weight_AVG_STER'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%")       
        """
        STD steering in 1 hour
        """
        i=0
        som_str = 0
        while (i < len(self.list_stdstr)):
            som_str += pow((self.list_stdstr[i] - avg_str), 2)
            i += 1
        std_str = pow((som_str/len(self.list_stdstr)), 1/2)
        if (std_str > data1['max_STD_STER_per_h']):
            perc_std_str = 0
        else:
            perc_std_str = (1 - (std_str/data1['max_STD_STER_per_h']))*data['weight_STD_STER']
        del self.list_stdstr[0:len(self.list_stdstr)]   
        self.list_hour.append(std_str)
        self.list_perc.append(str(int(perc_std_str)) +"/" +str(data['weight_STD_STER'])) 
        if  perc_std_str < data['weight_STD_STER']:
            p = (perc_std_str/data['weight_STD_STER'])*100
        else:
            p = 100
        self.list_p.append(str(int(p)) +"%") 
       
        
        som_evaluation = perc_crush +  perc_space + perc_r_over + perc_l_over + perc_avg_t + perc_std_t + perc_avg_d + perc_std_d + perc_avg_l + perc_lane + perc_hazardous + perc_avg_acc + perc_std_acc + perc_avg_dec + perc_std_dec + perc_avg_str + perc_std_str
        #print("The policy evaluation is: ", som_evaluation, " out of ", totWeight)
        return som_evaluation, totWeight

    def evaluateEP(self, list_evaluate):
        """Method that evaluate the Policy"""
        data = Json.config_json_weight(self)        
        data2 = Json.config_json_max_ep(self)
        perc_crush = 0.0  
        perc_space = 0.0 
        perc_r_over = 0.0   
        perc_l_over = 0.0   
        perc_avg_t = 0.0   
        perc_std_t = 0.0   
        perc_avg_d = 0.0  
        perc_std_d = 0.0  
        perc_avg_l = 0.0  
        perc_lane = 0.0  
        perc_hazardous = 0.0 
        perc_avg_acc = 0.0 
        perc_std_acc = 0.0   
        perc_avg_dec = 0.0   
        perc_std_dec = 0.0   
        perc_avg_str = 0.0  
        perc_std_str = 0.0 

        totWeight = data['weight_crush'] + data['weight_km'] + data['weight_right_overtake'] + data['weight_left_overtake'] + data['weight_AVG_TTC'] + data['weight_STD_TTC'] + data['weight_AVG_DSL'] + data['weight_STD_DSL'] + data['weight_AVG_LOAF'] + data['weight_number_lane_change'] + data['weight_hazardous'] + data['weight_AVG_ACC'] + data['weight_STD_ACC'] + data['weight_AVG_DEC'] + data['weight_STD_DEC'] + data['weight_AVG_STER'] + data['weight_STD_STER'] 
        h = 18*len(self.lista)
        """
        Crush in 1 hour
        """
        i= 0 + h 
        if (self.list_evaluate[i]) != data2['max_crush']:
            perc_crush = 0
        else:
            perc_crush = data['weight_crush']
        """
        Km in 1 hour
        """
        i=1 + h
        if (self.list_evaluate[i] > data2['max_km']):
            perc_space = data['weight_km']
        else:
            perc_space = (self.list_evaluate[i]/data2['max_km'])*data['weight_km']   
        """
        Right overtake in 1 hour
        """
        i=2 + h
        if (self.list_evaluate[i] > data2['max_right_overtake']):
            perc_r_over = 0 
        else:
            perc_r_over = (1 - (self.list_evaluate[i]/data2['max_right_overtake']))*data['weight_right_overtake']
        """
        Left overtake in 1 hour
        """
        i=3 + h
        if (self.list_evaluate[i] > data2['max_left_overtake']):
            perc_l_over = data['weight_left_overtake'] 
        else:
            perc_l_over = (self.list_evaluate[i]/data2['max_left_overtake'])*data['weight_left_overtake']  
        """
        AVG TTC in 1 hour
        """
        i=4 + h
        if (self.list_evaluate[i] > data2['max_AVG_TTC']):
            perc_avg_t = data['weight_AVG_TTC']
        else:
            perc_avg_t = (self.list_evaluate[i]/data2['max_AVG_TTC'])*data['weight_AVG_TTC']
        """
        STD TTC in 1 hour
        """
        i=5 + h
        if (self.list_evaluate[i] > data2['max_STD_TTC']):
            perc_std_t = 0 
        else: 
            perc_std_t = (1 - (self.list_evaluate[i]/data2['max_STD_TTC']))*data['weight_STD_TTC']
        """
        AVG delta speed limit in 1 hour
        """
        i=6 + h
        if (self.list_evaluate[i] > data2['max_AVG_DSL']):
            perc_avg_d = 0
        else: 
            perc_avg_d = (1 - (self.list_evaluate[i]/data2['max_AVG_DSL']))*data['weight_AVG_DSL']    
        """
        STD delta speed limit in 1 hour
        """
        i=7 + h
        if (self.list_evaluate[i] > data2['max_STD_DSL']):
            perc_std_d = 0 
        else:
            perc_std_d = (1 - (self.list_evaluate[i]/data2['max_STD_DSL']))*data['weight_STD_DSL'] 
        """
        AVG lane occupacy function
        """
        i=8 + h
        perc_avg_l = (self.list_evaluate[i]/data2['max_AVG_LOAF'])*data['weight_AVG_LOAF']
        """
        Lanes change in 1 hour
        """
        i=9 + h
        if (self.list_evaluate[i] > data2['max_number_lane_change']):
            perc_lane = 0
        else:
            perc_lane = data['weight_number_lane_change'] 
        """
        Hazardous lane change in 1 h
        """
        i=10 + h
        if (self.list_evaluate[i] > data2['max_hazardous']):
            perc_hazardous = 0
        else:
            perc_hazardous = (1- (self.list_evaluate[i]/data2['max_hazardous']))*data['weight_hazardous']  
        
        """
        AVG acceleration in 1 h
        """
        i=11 + h
        if (self.list_evaluate[i] > data2['max_AVG_ACC']):
            perc_avg_acc = 0 
        else:
            perc_avg_acc = (1 - (self.list_evaluate[i]/data2['max_AVG_ACC']))*data['weight_AVG_ACC'] 
        """
        STD acceleration in 1 h
        """
        i=12 + h
        if (self.list_evaluate[i] > data2['max_STD_ACC']):
            perc_std_acc = 0 
        else: 
            perc_std_acc = (1 - (self.list_evaluate[i]/data2['max_STD_ACC']))*data['weight_STD_ACC']
        """
        AVG deceleration in 1 h
        """
        i=13 + h
        if (self.list_evaluate[i] > data2['max_AVG_DEC']):
            perc_avg_acc = 0  
        else:
            perc_avg_dec = (1- (self.list_evaluate[i]/data2['max_AVG_DEC']))*data['weight_AVG_DEC']  
        """
        STD deceleration 1 hour
        """
        i=14 + h
        if (self.list_evaluate[i] > data2['max_STD_DEC']):
            perc_std_dec = 0 
        else:
            perc_std_dec = (1 - (self.list_evaluate[i]/data2['max_STD_DEC']))*data['weight_STD_DEC'] 
        """
        AVG steering in 1 hour
        """       
        i=15 + h
        if (self.list_evaluate[i] > data2['max_AVG_STER']):
            perc_avg_str = 0
        else: 
            perc_avg_str = (1 - (self.list_evaluate[i]/data2['max_AVG_STER']))*data['weight_AVG_STER']      
        """
        STD steering in 1 hour
        """
        i=16 + h
        if (self.list_evaluate[i] > data2['max_STD_STER']):
            perc_std_str = 0
        else:
            perc_std_str = (1 - (self.list_evaluate[i]/data2['max_STD_STER']))*data['weight_STD_STER']
        
        
        som_ep = perc_crush +  perc_space + perc_r_over + perc_l_over + perc_avg_t + perc_std_t + perc_avg_d + perc_std_d + perc_avg_l + perc_lane + perc_hazardous + perc_avg_acc + perc_std_acc + perc_avg_dec + perc_std_dec + perc_avg_str + perc_std_str
        #print("The policy evaluation is: ", som_evaluation, " out of ", totWeight)
        return som_ep

            